from django.http import HttpResponse
from django.shortcuts import render, redirect,get_object_or_404
from .models import *
from django.http import HttpResponse, HttpResponseRedirect
from .models import InsuranceEnquiry
import datetime
from django.core.exceptions import ValidationError
from datetime import datetime
from .models import LoanEnquiry, Document
from django.core.exceptions import ValidationError
from openpyxl import Workbook
from django.contrib import messages
from .models import LoanEnquiry
from .forms import InsuranceEnquiryForm
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from .models import Users
from .models import Loan
from .forms import UserForm
from django.contrib.auth.forms import AuthenticationForm, PasswordResetForm
from django.contrib.auth import login as auth_login
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User 
from django.contrib.auth import authenticate,login
from django.shortcuts import render, redirect
from django.contrib.auth.forms import PasswordResetForm
from django.contrib import messages
import plotly.express as px 
from datetime import datetime, date

def success(request):
     return render(request,'success.html')

def untitle(request):
     return render(request,'untitle.html')


def Logout(request):
    logout(request)
    return redirect('login')



# def HomePage(request):
#      return render (request, 'index.html')
from django.shortcuts import render
import plotly.graph_objects as go

from django.db.models import Sum
from django.shortcuts import render
import plotly.graph_objects as go
from .models import PolicyIssue  # Import your model here
from .models import PolicyIssue, Loan, InsuranceEnquiry


# def HomePage(request):
#      return render (request, 'index.html')
from django.shortcuts import render
import plotly.graph_objects as go

from django.db.models import Sum
from django.shortcuts import render
import plotly.graph_objects as go
from .models import PolicyIssue  # Import your model here
from .models import PolicyIssue, Loan, InsuranceEnquiry


def HomePage(request):
    # Retrieve data from your database
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')



    queryset = PolicyIssue.objects.all()  # Replace PolicyIssue with your actual model
    queryset1 = Loan.objects.all()
    inquiries = InsuranceEnquiry.objects.all()

    if start_date and end_date:
        queryset = queryset.filter(date__range=[start_date, end_date])
    # Calculate metrics (example calculation)
       # Calculate metrics
    total_p_number = queryset.count()
    total_payout_amount = queryset.aggregate(total_payout_amount=Sum('PayoutAmount'))['total_payout_amount']

    total_premium = queryset.aggregate(total_premium=Sum('Premium'))['total_premium']
    total_odnet_premium = queryset.aggregate(total_odnet_premium=Sum('odNetPremium'))['total_odnet_premium']
    total_commission_percentage = queryset.aggregate(total_commission_percentage=Sum('commissionPercentage'))['total_commission_percentage']
    total_profit = queryset.aggregate(total_profit=Sum('profitResult'))['total_profit']
    total_net_profit = queryset.aggregate(total_net_profit=Sum('netProfitResult'))['total_net_profit']
    total_profit_after_tds = queryset.aggregate(total_profit_after_tds=Sum('profitAfterTDSResult'))['total_profit_after_tds']
    # Get the total count of unique names
    total_unique_names = inquiries.values('name').distinct().count()
    # Calculate the total number of loan inquiries
    total_name = queryset1.count() 
    # Create Plotly figure for each metric

    fig_payout_amount = go.Figure(go.Indicator(
    mode="number",
    value=total_payout_amount,
    title="Total Payout Amount",
    number={'font': {'size': 40}}  # Adjust the font size of the number
    ))
    fig_payout_amount.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=20)  # Adjust the font size of the title
)
    fig_profit = go.Figure(go.Indicator(
        mode="number",
        value=total_profit,
        title="Total Profit",
        number={'font':{'size': 40}}
    ))
    fig_profit.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)
    

    fig_net_profit = go.Figure(go.Indicator(
        mode="number",
        value=total_net_profit,
        title="Total Net Profit",
        number={'font':{'size': 40}}
    ))
    fig_net_profit.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)
    

    fig_profit_after_tds = go.Figure(go.Indicator(
        mode="number",
        value=total_profit_after_tds,
        title="Total Profit After TDS",
        number={'font':{'size': 40}}
        
    ))
    fig_profit_after_tds.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)

    # Create Plotly figures for each metric
    fig_customer = go.Figure(go.Indicator(
        mode="number",
        value=total_p_number,
        title="Total PolicyIssue",
        number={'font':{'size': 40}}
    ))
    fig_customer.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)

    fig_premium = go.Figure(go.Indicator(
        mode="number",
        value=total_premium,
        title="Total Premium",
        number={'font':{'size': 40}}
    ))
    fig_premium.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)

    fig_odnet_premium = go.Figure(go.Indicator(
        mode="number",
        value=total_odnet_premium,
        title="Total OD/Net Premium",
        number={'font':{'size': 40}}
    ))
    fig_odnet_premium.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)

    fig_commission_percentage = go.Figure(go.Indicator(
        mode="number",
        value=total_commission_percentage,
        title="Total Commission",
        number={'font':{'size': 40}}
    ))
    fig_commission_percentage.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)
    
    # Create a Plotly figure for the total count of unique names
    fig_total_unique_names = go.Figure(go.Indicator(
        mode="number",
        value=total_unique_names,
        title="Total InsuranceEnquiry",
        number={'font':{'size': 40}}
    ))
    fig_total_unique_names.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)
    
     # Create Plotly figures for each metric
    # Create Plotly figures for each metric
    fig_total_name = go.Figure(go.Indicator(
        mode="number",
        value=total_name,
        title="Total LoanEnquiry",
        number={'font':{'size': 40}}
    ))
    fig_total_name.update_layout(
    width=200,  # Set the width of the Indicator
    height=100,  # Set the height of the Indicator
    font=dict(size=15)  # Adjust the font size of the title
)
     # Use current month as default value
    default_start_date = None
    default_end_date = None

    if start_date is None or start_date == 'default_start_date':
        start_date = default_start_date
    if end_date is None or end_date == 'default_end_date':
        end_date = default_end_date

    if start_date != 'default_start_date' and end_date != 'default_end_date' and start_date is not None and end_date is not None:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    df = fetch_data(request)

    if start_date != 'default_start_date' and end_date != 'default_end_date' and start_date is not None and end_date is not None:
        df = df[(df['date'] >= start_date) & (df['date'] <= end_date)]

    df['date'] = pd.to_datetime(df['date'])


    # Grouping data by insurance company and summing premiums
    df_grouped = df.groupby('I_company')['Premium'].sum().reset_index()

    # Sorting the companies based on total premiums
    df_sorted = df_grouped.sort_values(by='Premium', ascending=False)

    # Selecting the top 5 companies
    top_5_companies = df_sorted.head(5)

    # Plotting the bar chart
    fig_bar = px.bar(top_5_companies, x="I_company", y="Premium", title="<b>Top 5 Insurance Companies by Premium")

    fig_bar.update_layout(
    # autosize=False,
    width=450,
    height=500,
    margin=dict(
        l=40,
        r=40,
        b=60,
        t=80,
        pad=3
    ),
    plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
    paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
)
    fig_bar.update_xaxes(title_text='Insurance Company')
    fig_bar.update_yaxes(title_text='Premium')
    # Extracting month from the date column
    df['Month'] = df['date'].dt.month_name()

    # Define the order of months
    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

    # Convert 'Month' column to categorical with desired order
    df['Month'] = pd.Categorical(df['Month'], categories=month_order, ordered=True)

    # Grouping by month and summing up the values
    df1 = df.groupby(by="Month")[["netProfitResult", "profitAfterTDSResult", "odNetPremium"]].sum().rename(columns={"netProfitResult": "Net Profit", "profitAfterTDSResult": "Profit After TDS", "odNetPremium": "OD Net Premium"})

    # Resetting index to make 'Month' a column again
    df1.reset_index(inplace=True)

    # Plotting the line chart with month on the x-axis
    fig_line = px.line(df1, x='Month', y=['Net Profit','Profit After TDS','OD Net Premium'], 
                title='<b>Net Profit Over Time with Additional Metrics</b>')

    fig_line.update_layout(
        # autosize=False,
        width=600,
        height=500,
        margin=dict(
            l=50,
            r=50,
            b=100,
            t=100,
            pad=4
        ),
        plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
        paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
    )

     # Plotting the donut chart using Plotly Graph Objects
    fig_donut_2 = go.Figure(data=[go.Pie(labels=df['business_type'], values=df['netProfitResult'], hole=0.5)])
    fig_donut_2.update_layout(title='Net profit by business_type')

    fig_donut_2.update_layout(
    title='<b>Net profit by insurance type',
    width=400,
    height=400,
    margin=dict(l=50, r=50, b=100, t=100, pad=4),
    plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
    paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
    legend=dict(
        orientation="h",  # Horizontal legend
        yanchor="bottom",  # Anchor legend to the bottom
        y=1.02,  # Adjust the position vertically
        xanchor="right",  # Anchor legend to the right
        x=1  # Adjust the position horizontally
    )
)

    # Plotting the donut chart using Plotly Graph Objects
    fig_donut = go.Figure(data=[go.Pie(labels=df['insurance_type'], values=df['netProfitResult'], hole=0.5)])
    fig_donut.update_layout(title="Net profit by insurance type")

    fig_donut.update_layout(
    title="<b>Net profit by insurance type",
    width=400,
    height=400,
    margin=dict(l=50, r=50, b=100, t=100, pad=6),
    plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
    paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
    legend=dict(
        orientation="h",  # Horizontal legend
        yanchor="bottom",  # Anchor legend to the bottom
        y=1.02,  # Adjust the position vertically
        xanchor="right",  # Anchor legend to the right
        x=1  # Adjust the position horizontally
    )
)

    # fig_funnel = px.funnel(df, x='Location', y='Premium')
    df_grouped = df.groupby('Location')['Premium'].sum().reset_index()
    df_sorted = df_grouped.sort_values(by='Premium', ascending=False)

    # Selecting the top 5 companies
    top_5_companies = df_sorted.head(5)

    # Plotting the bar chart
    fig_funnel = px.funnel(top_5_companies, x="Premium", y="Location", title="<b>Location by Premium",orientation='h')

    fig_funnel.update_layout(
    # autosize=False,
    width=390,
    height=500,
    margin=dict(
        l=50,
        r=50,
        b=100,
        t=100,
        pad=4
    ),
    plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
    paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
)


    # Convert Plotly figures to HTML
    plot_html_payout_amount = fig_payout_amount.to_html(full_html=False)
    plot_html_profit = fig_profit.to_html(full_html=False)
    plot_html_net_profit = fig_net_profit.to_html(full_html=False)
    plot_html_profit_after_tds = fig_profit_after_tds.to_html(full_html=False)
    plot_html_p_number = fig_customer.to_html(full_html=False)
    plot_html_premium = fig_premium.to_html(full_html=False)
    plot_html_odnet_premium = fig_odnet_premium.to_html(full_html=False)
    plot_html_commission_percentage = fig_commission_percentage.to_html(full_html=False)
    plot_html_total_unique_names = fig_total_unique_names.to_html(full_html=False)
    plot_html_total_name = fig_total_name.to_html(full_html=False)
    # Convert the Plotly figures to HTML
    plot_html_bar = fig_bar.to_html(full_html=False)
    plot_html_line = fig_line.to_html(full_html=False)
    plot_html_donut_2 = fig_donut_2.to_html(full_html=False)
    plot_html_donut = fig_donut.to_html(full_html=False)
    plot_html_funnel = fig_funnel.to_html(full_html=False)


    
       # Pass the HTML to the template for rendering
    return render(request, 'index.html',{'plot_html_payout_amount': plot_html_payout_amount,
                                              'plot_html_profit': plot_html_profit,
                                              'plot_html_net_profit': plot_html_net_profit,
                                              'plot_html_profit_after_tds': plot_html_profit_after_tds,
                                              'plot_html_p_number': plot_html_p_number,
                                              'plot_html_premium': plot_html_premium,
                                              'plot_html_odnet_premium': plot_html_odnet_premium,
                                              'plot_html_commission_percentage': plot_html_commission_percentage,
                                              'plot_html_total_unique_names': plot_html_total_unique_names,
                                              'plot_html_total_name': plot_html_total_name,
                                              'plot_html_bar': plot_html_bar, 
                                              'plot_html_line': plot_html_line, 
                                              'plot_html_donut_2': plot_html_donut_2, 
                                              'plot_html_donut': plot_html_donut,
                                              'plot_html_funnel':plot_html_funnel})

def logout(request):
     return redirect ('login')

from django.shortcuts import render, redirect
from .forms import InsuranceEnquiryForm
from .models import InsuranceEnquiry

def enquiry2(request):
    if request.method == 'POST':
        form = InsuranceEnquiryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            # Redirect to success page or do something else
            return redirect('enquiry2')  # Change 'enquiry2' to the URL name of your success page
        else:
            # If form is invalid, render the form again with error messages
            return render(request, 'enquiry2.html', {'form': form})
    else:  # Handling GET request
        form = InsuranceEnquiryForm()  # Initialize an empty form
        enquiries = InsuranceEnquiry.objects.all()
        return render(request, 'enquiry2.html', {'form': form, 'enquiries': enquiries})


# def enquiry2(request):
#      form = InsuranceEnquiryForm()  # Define form variable here
#      if request.method == 'POST':
#         # Extract data from the form
#         name = request.POST.get('name')
#         number = request.POST.get('number')
#         email = request.POST.get('email')
#         vehicle_number = request.POST.get('vehicle_number')
#         rc_book_radio = request.POST.get('rc_book_radio')
#         previous_policy_radio = request.POST.get('previous_policy_radio')
#         end_date = request.POST.get('end_date')

#         # Validate end date format
#         if end_date:
#             try:
#                 # Try to convert end date to a valid date object
#                 from datetime import datetime
#                 datetime.strptime(end_date, '%Y-%m-%d')
#             except ValueError:
#                 # If end date is not in the correct format, handle the error
#                 return render(request, 'form.html', {'error_message': 'End date must be in YYYY-MM-DD format'})

#         # Extract images from the request.FILES dictionary
#         rc_book_image = request.FILES.get('rc_book_image')
#         previous_policy_image = request.FILES.get('previous_policy_image')

#         # Save data to the model
#         enquiry = InsuranceEnquiry(
#             name=name,
#             number=number,
#             email=email,
#             vehicle_number=vehicle_number,
#             rc_book_radio=rc_book_radio,
#             previous_policy_radio=previous_policy_radio,
#             end_date=end_date,
#             rc_book_image=rc_book_image,  # Save RC book image
#             previous_policy_image=previous_policy_image  # Save previous policy image
#         )
#         enquiry.save()

#         # Redirect to success page or do something else
#         return redirect('enquiry2')  # Change 'success' to the URL name of your success page
    
#     # Fetch data from the database
#      enquiries = InsuranceEnquiry.objects.all()
    
#     # Pass data to the template for rendering
#      return render(request, 'enquiry2.html', {'form': form, 'enquiries': enquiries})

def SignupPage(request):
    if request.method=='POST':
        username=request.POST.get('username')
        email=request.POST.get('email')
        password=request.POST.get('password')
        
        my_user=User.objects.create_user(username,email,password)
        my_user.save()
        return redirect('login')
       
    return render (request, 'signup.html')

def LoginPage(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            # Process form data, authenticate user, etc.
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Check if the user should be redirected to the home page
                next_url = request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                else:
                    return redirect('home')  # Redirect to home page after successful login
            else:
                return HttpResponse("Username or Password is incorrect!!!")
    else:
        form = AuthenticationForm()  # Instantiate form object if request method is not 'POST'

    return render(request, 'login.html', {'form': form})

def forgot_password(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            form.save(request=request)
            messages.success(request, 'An email has been sent with instructions to reset your password.')
            return redirect('login')  # Redirect to the login page after password reset request
    else:
        form = PasswordResetForm()
        
    return render(request, 'forgotpwd.html', {'form': form})



# from django.contrib.auth import authenticate, login as auth_login
# from django.contrib.auth.forms import UserCreationForm, AuthenticationForm, PasswordResetForm
# from django.shortcuts import render, redirect
# from django.contrib.auth.models import User  # Django ke built-in User model ka import karein

# # @login_required(login_url="login/")
# def home(request):
#     return render(request, 'index.html')

# def register(request):
#     if request.method == 'POST':
#         # Extract data from the POST request
#         username = request.POST.get('username')
#         email = request.POST.get('email')
#         password = request.POST.get('password')

        
#         # Create a new user object and save it to the database
#         user = User.objects.create_user(
#             username=username,
#             email=email,
#             password=password
#         )
#         user.save()
#         # Redirect to the login page
#         return redirect('login')  # Assuming 'login' is the name of your login URL pattern
            
    
#     # Render the registration form if it's a GET request
#     return render(request, 'register.html')

# from django.contrib.auth import authenticate, login
# from django.shortcuts import render, redirect

# def login_view(request):
#     if request.method == 'POST':
#         username = request.POST('username')
#         password = request.POST('password')
#         user = authenticate(request, username=username, password=password)
#         if user is not None:
#             login(request, user)
#             # Redirect to a success page.
#             return redirect('index')
#         else:
#             # Return an 'invalid login' error message.
#             return render(request, 'login.html', {'error_message': 'Invalid login'})
#     else:
#         return render(request, 'login.html')



# def register(request):
#     if request.method == 'POST':
#         form = UserCreationForm(request.POST)
#         if form.is_valid():
#             form.save()
#             # Automatically login user after 
#             # registration
#             email = form.cleaned_data.get('email')
#             password = form.cleaned_data.get('password1')
#             user = authenticate(username=email, password=password)
#             if user is not None:
#                 auth_login(request, user)
#                 return redirect('login')  # Redirect to home page after successful registration and login
#     else:
#         form = UserCreationForm()
#     return render(request, 'register.html', {'form': form})

# def login(request):
#     if request.method == 'POST':
#         form = AuthenticationForm(data=request.POST)
#         if form.is_valid():
#             user = form.get_user()
#             auth_login(request, user)
#             return redirect('home')  # Redirect to home page after successful login
#     else:
#         form = AuthenticationForm()
#     return render(request, 'login.html', {'form': form})

# def forgot_password(request):
#     if request.method == 'POST':
#         form = PasswordResetForm(request.POST)
#         if form.is_valid():
#             # Handle forgot password logic here
#             pass
#     else:
#         form = PasswordResetForm()
#     return render(request, 'forgotpwd.html', {'form': form})


# def login_view(request):
#     if request.method == 'POST':
#         # Extract username and password from the POST request
#         username = request.POST.get('username')
#         password = request.POST.get('password')
#         try:
#             # Retrieve user object from the database based on email and password
#             user = Users.objects.get(email=username, password=password)
#             # If the user is found, set the user_id in the session and redirect to home page
#             request.session['user'] = username
#             return render(request,'index.html')  # Assuming the URL name for home page is 'home'
#         except Users.DoesNotExist:
#             # Handle the case where the user does not exist or password is incorrect
#             return HttpResponse("Invalid credentials.")
#     else:
#         # Render the login form if it's a GET request
#         return render(request, 'login.html')
    
# def register(request):
#     if request.method == 'POST':
#         # Extract data from the POST request
#         name = request.POST.get('name')
#         email = request.POST.get('email')
#         password = request.POST.get('password')

#         # Check if all required fields are present
#         if name and email and password:
#             try:
#                 # Create a new user object and save it to the database
#                 user = Users.objects.create(
#                     name=name,
#                     email=email,
#                     password=password
#                 )
#                 # Redirect to the login page
#                 return redirect('login')  # Assuming 'login' is the name of your login URL pattern
#             except IntegrityError:
#                 # Handle the case where the email address already exists
#                 return render(request, 'register.html', {'error_message': 'This email address is already registered.'})
#             except Exception as e:
#                 # Handle other exceptions
#                 print(e)  # Print the error message for debugging purposes
#                 return render(request, 'register.html', {'error_message': 'An error occurred while processing your request.'})
#         else:
#             # Handle the case where required fields are missing
#             return render(request, 'register.html', {'error_message': 'All fields are required.'})
    
#     # Render the registration form if it's a GET request
#     return render(request, 'register.html')





# # def register(request):
# #     if request.method == 'POST':
# #         # Extract data from the POST request
# #         name = request.POST.get('name')
# #         email = request.POST.get('email')
# #         password = request.POST.get('password')

# #         # Check if all required fields are present
# #         if name and email and password:
# #             try:
# #                 # Create a new user object and save it to the database
# #                 user = Users.objects.create(
# #                     name=name,
# #                     email=email,
# #                     password=password
# #                 )
# #                 # Redirect to the login page
# #                 return redirect('login/')  # Assuming 'login' is the name of your login URL pattern
# #             except IntegrityError:
# #                 # Handle the case where the email address already exists
# #                 return render(request, 'register.html', {'error_message': 'This email address is already registered.'})
# #         else:
# #             # Handle the case where required fields are missing
# #             return render(request, 'register.html', {'error_message': 'All fields are required.'})
    
# #     # Render the registration form if it's a GET request
# #     return render(request, 'register.html')

# def forgotPwd(request):
#     if request.method == 'POST':
#         # Extract common data
#         email = request.POST.get('email')
#         password = request.POST.get('password')

#         # user = authenticate(request, username = username, password = password)

#         # if user is not None:
#         #     login(request, user)
#         try:
#             user = Users.objects.get(email=email)
#             user.password = password
#             user.save()
            
#             return render(request, 'base.html')
#         except:
#             return render(request, 'forgotpwd.html')

#     return render(request, 'forgotpwd.html')


from .models import InsuranceEnquiry
from .forms import InsuranceEnquiryForm

from .models import InsuranceEnquiry
from .forms import InsuranceEnquiryForm

def insurance_enquiry(request):
    form = InsuranceEnquiryForm()  # Define form variable here

    if request.method == 'POST':
        # Extract data from the form
        name = request.POST.get('name')
        number = request.POST.get('number')
        email = request.POST.get('email')
        vehicle_number = request.POST.get('vehicle_number')
        rc_book_radio = request.POST.get('rc_book_radio')
        previous_policy_radio = request.POST.get('previous_policy_radio')
        end_date = request.POST.get('end_date')

        # Validate end date format
        if end_date:
            try:
                # Try to convert end date to a valid date object
                from datetime import datetime
                datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                # If end date is not in the correct format, handle the error
                return render(request, 'form.html', {'error_message': 'End date must be in YYYY-MM-DD format'})

        # Extract images from the request.FILES dictionary
        rc_book_image = request.FILES.get('rc_book_image')
        previous_policy_image = request.FILES.get('previous_policy_image')

        # Save data to the model
        enquiry = InsuranceEnquiry(
            name=name,
            number=number,
            email=email,
            vehicle_number=vehicle_number,
            rc_book_radio=rc_book_radio,
            previous_policy_radio=previous_policy_radio,
            end_date=end_date,
            rc_book_image=rc_book_image,  # Save RC book image
            previous_policy_image=previous_policy_image  # Save previous policy image
        )
        enquiry.save()

        # Redirect to success page or do something else
        return redirect('insurance_enquiry')  # Change 'success' to the URL name of your success page
    
    # Fetch data from the database
    enquiries = InsuranceEnquiry.objects.all()
    
    # Pass data to the template for rendering
    return render(request, 'form.html', {'form': form, 'enquiries': enquiries})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import InsuranceEnquiry
from .forms import InsuranceEnquiryForm
from django.urls import reverse

@csrf_exempt
def update_insurance_enquiry(request, record_id):
    enquiries = InsuranceEnquiry.objects.all()
    enquiry = InsuranceEnquiry.objects.get(pk=record_id)
    
    if request.method == 'POST':
        form1 = InsuranceEnquiryForm(request.POST, request.FILES, instance=enquiry)
        if form1.is_valid():
            form1.save()
            return HttpResponseRedirect(reverse('insurance_enquiry'))  # Redirect on successful form submission
    else:
        form1 = InsuranceEnquiryForm(instance=enquiry)
    
    return render(request, 'form.html', {'form': form1,'enquiries':enquiries})


# def edit_view(request, record_id):
#     record = get_object_or_404(InsuranceEnquiry, pk=record_id)
    
#     if request.method == 'POST':
#         # Retrieve the data from the request
#         name = request.POST.get('name')
#         number = request.POST.get('number')
#         email = request.POST.get('email')
#         vehicle_number = request.POST.get('vehicle_number')
#         rc_book_radio = request.POST.get('rc_book_radio')
#         previous_policy_radio = request.POST.get('previous_policy_radio')
#         end_date_str = request.POST.get('end_date')

#         # Update the record fields
#         record.name = name
#         record.number = number
#         record.email = email
#         record.vehicle_number = vehicle_number

#         # Handle RC Book data
#         if rc_book_radio == 'yes':
#             record.rc_book_image = request.FILES.get('rc_book_image')
#         else:
            
#             record.rc_book_image = None
            
#         # Handle Previous Policy data
#         if previous_policy_radio == 'yes':
#             record.previous_policy_image = request.FILES.get('previous_policy_image')
#             try:
#                 record.end_date = parse_date(end_date_str)
#             except ValidationError:
#                 # Handle the case where the end date is not in the correct format
#                 pass
#         else:
#             record.previous_policy_image = None
#             record.end_date = None
        
#         # Save the updated record
#         record.save()

#         # Redirect to a success page or any other page after successful update
#         return redirect('update_enquiry', record_id=record.id)
#     else:
#         return render(request, 'editform.html', {'record': record})

# from django.shortcuts import render, redirect, get_object_or_404
# from django.core.exceptions import ValidationError
# from django.utils.dateparse import parse_date
# from .models import InsuranceEnquiry

# def edit_view(request, record_id):
#     record = get_object_or_404(InsuranceEnquiry, pk=record_id)
#     # Your existing view code here:
#     if request.method == 'POST':
#         # Retrieve the data from the request
#         name = request.POST.get('name')
#         number = request.POST.get('number')
#         email = request.POST.get('email')
#         vehicle_number = request.POST.get('vehicle_number')
#         rc_book_radio = request.POST.get('rc_book_radio')
#         previous_policy_radio = request.POST.get('previous_policy_radio')
#         end_date_str = request.POST.get('end_date')

#         # Find the record based on vehicle_number
#         record = get_object_or_404(InsuranceEnquiry, vehicle_number=vehicle_number)

#         # Update the record fields
#         record.name = name
#         record.number = number
#         record.email = email

#         # Handle RC Book data
#         if rc_book_radio == 'Yes':
#             record.rc_book_image = request.FILES.get('rc_book_image')

#         # Handle Previous Policy data
#         if previous_policy_radio == 'yes':
#             record.previous_policy_image = request.FILES.get('previous_policy_image')
#             # Handle end_date
#             try:
#                 record.end_date = parse_date(end_date_str)
#             except ValidationError:
#                 # Handle the case where the end date is not in the correct format
#                 pass

#         # Save the updated record
#         record.save()

#         # Redirect to a success page or any other page after successful update
#         return redirect('update_enquiry',record_id=record.id)
#     else:
#         # Render the form template
#         # Assuming 'record' is the object you want to edit
#         return render(request, 'editform.html', {'record': record})

from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
from .models import InsuranceEnquiry

def edit_view(request, record_id):
    record = get_object_or_404(InsuranceEnquiry, pk=record_id)
    
    if request.method == 'POST':
        # Retrieve the data from the request
        name = request.POST.get('name')
        number = request.POST.get('number')
        email = request.POST.get('email')
        vehicle_number = request.POST.get('vehicle_number')
        rc_book_radio = request.POST.get('rc_book_radio')
        previous_policy_radio = request.POST.get('previous_policy_radio')
        end_date_str = request.POST.get('end_date')

        # Find the record based on vehicle_number
        record = get_object_or_404(InsuranceEnquiry, vehicle_number=vehicle_number)

        # Update the record fields
        record.name = name
        record.number = number
        record.email = email

        # Handle RC Book data
        if rc_book_radio == 'Yes':
            rc_book_image = request.FILES.get('rc_book_image')
            if rc_book_image:
                record.rc_book_image = rc_book_image  # Set RC Book Image if submitted
            record.rc_book_radio = 'Yes'  # Ensure RC Book is set to Yes
        else:
            record.rc_book_image = None  # Clear the RC Book Image if RC Book is No
            record.rc_book_radio = 'No'   # Set RC Book to No if RC Book is No

        # Handle Previous Policy data
        if previous_policy_radio == 'yes':
            previous_policy_image = request.FILES.get('previous_policy_image')
            if previous_policy_image:
                record.previous_policy_image = previous_policy_image  # Set Previous Policy Image if submitted
            record.previous_policy_radio = 'Yes'  # Ensure Previous Policy is set to Yes
            # Handle end_date
            try:
                record.end_date = parse_date(end_date_str)
            except ValidationError:
                # Handle the case where the end date is not in the correct format
                pass
        else:
            record.previous_policy_image = None  # Clear the Previous Policy Image if Previous Policy is No
            record.previous_policy_radio = 'No'   # Set Previous Policy to No if Previous Policy is No
            record.end_date = None  # Clear the End Date if Previous Policy is No

        # Save the updated record
        record.save()

        # Redirect to a success page or any other page after successful update
        return redirect('update_enquiry', record_id=record.id)
    else:
        # Render the form template
        # Assuming 'record' is the object you want to edit
        return render(request, 'editform.html', {'record': record})



def update_enquiry(request, pk):
    enquiry = get_object_or_404(InsuranceEnquiry, pk=pk)
    form = InsuranceEnquiryForm(instance=enquiry)

    if request.method == 'POST':
        form = InsuranceEnquiryForm(request.POST, request.FILES, instance=enquiry)
        if form.is_valid():
            form.save()
            return redirect('insurance_enquiry')

    return render(request, 'editform.html', {'form': form, 'enquiry': enquiry})

def delete_enquiry(request, enquiry_id):
    enquiry = get_object_or_404(InsuranceEnquiry, pk=enquiry_id)
    if request.method == 'POST':
        enquiry.delete()
        return redirect('insurance_enquiry')  # Redirect to the insurance enquiry page
    else:
        # Handle GET request method if needed
        pass

# def loan_enquiry(request):
#     if request.method == 'POST':
#         # Extract data from the form
#         name = request.POST.get('name')
#         number = request.POST.get('number')
#         email = request.POST.get('email')
#         rc_book_radio = request.POST.get('rc_book_radio')
#         rc_book_image = request.FILES.get('rc_book_image')
#         documents = request.FILES.getlist('documents')  # Handle multiple file uploads
        
#         # Save data to the model
#         enquiry = LoanEnquiry(
#             name=name,
#             number=number,
#             email=email,
#             rc_book_radio=rc_book_radio,
#             rc_book_image=rc_book_image,
#         )
#         enquiry.save()
        
#         # Save document files
#         for document in documents:
#             doc = Document(document=document)
#             doc.save()
#             enquiry.documents.add(doc)
        
#         # Redirect to success page or do something else
#         return redirect('success')  # Adjust the URL name according to your project
    
#     # Fetch data from the database
#     data = LoanEnquiry.objects.all()
    
#     # Pass data to the template for rendering
#     return render(request, 'loan.html', {'data': data})

from django.shortcuts import render, redirect, get_object_or_404
from .forms import LoanEnquiryForm  # YourFormClass ko aapko apne form class ke saath replace karna hoga
from .models import Loan  # YourModel ko aapko apne model ke saath replace karna hoga

# def edit_loan(request, record_id):
#     print("oooooooooooopppppppppppppppp")
#     loan_record = get_object_or_404(Loan, pk=record_id)
#     loan_instance = Loan.objects.get(pk=record_id)
#     if request.method == 'POST':
#         name = request.POST.get('name')
#         number = request.POST.get('number')
#         email = request.POST.get('email')
#         rc_book_radio = request.POST.get('rc_book_radio')
        
#         # Update the loan instance
#         loan_instance.name = name
#         loan_instance.number = number
#         loan_instance.email = email
#         loan_instance.rc_book_radio = rc_book_radio
        
#         # Handle file upload if any
#         rc_book_image = request.FILES.get('rc_book_image')
#         if rc_book_image:
#             loan_instance.rc_book_image = rc_book_image
        
#         # Save the updated instance
#         loan_instance.save()
        
#         return redirect('loan')  # Replace 'success_page' with your actual URL name
#     return render(request, 'loanedit.html', {'form': loan_instance, 'record_id': record_id, 'documents': loan_record.documents})
    
# views.py

from django.shortcuts import render, redirect, get_object_or_404
from .models import Loan
from .forms import LoanEnquiryForm  # Assuming you have a form for LoanEnquiry
def loan(request):
    if request.method == 'POST':
        # Extract data from the form
        name = request.POST.get('name')
        number = request.POST.get('number')
        email = request.POST.get('email')
        remarks = request.POST.get('remarks')  # Make sure this matches the name attribute in the form
        rc_book_radio = request.POST.get('rc_book_radio')
        rc_book_image = request.FILES.get('rc_book_image')
        document1 = request.FILES.get('document1')
        document2 = request.FILES.get('document2')
        document3 = request.FILES.get('document3')
        document4 = request.FILES.get('document4')
        document5 = request.FILES.get('document5')

        # Convert rc_book_radio to boolean
        rc_book = True if rc_book_radio == 'yes' else False

        # Save data to the model
        loan = Loan(
            name=name,
            number=number,
            email=email,
            remarks=remarks,  # Make sure to assign the remarks value
            rc_book=rc_book,
            rc_book_image=rc_book_image,
            document1=document1,
            document2=document2,
            document3=document3,
            document4=document4,
            document5=document5,
        )
        loan.save()

        # Fetch all loan records
        loans = Loan.objects.all()
        # Redirect to success page or do something else
        return redirect('loan')  # Adjust the URL name according to your project

    # Fetch data from the database (if required)
    loans = Loan.objects.all()

    # Pass data to the template for rendering
    return render(request, 'loan.html', {'data': loans})


def edit_loan(request, record_id):
    loan_instance = get_object_or_404(Loan, pk=record_id)
    
    if request.method == 'POST':
        # Extract individual field values from the request
        name = request.POST.get('name')
        number = request.POST.get('number')
        email = request.POST.get('email')
        remarks = request.POST.get('remarks')
        rc_book_radio = request.POST.get('rc_book_radio')
        rc_book_image = request.FILES.get('rc_book_image')
        document1 = request.FILES.get('document1')
        document2 = request.FILES.get('document2')
        document3 = request.FILES.get('document3')
        document4 = request.FILES.get('document4')
        document5 = request.FILES.get('document5')

        # Update the loan_instance with new field values
        loan_instance.name = name
        loan_instance.number = number
        loan_instance.email = email
        loan_instance.remarks = remarks
        loan_instance.rc_book_radio = rc_book_radio

        # If RC Book Image is provided, save it
        if rc_book_image:
            loan_instance.rc_book_image = rc_book_image

        # Save the loan_instance to update the existing record
        loan_instance.save()

        # Save documents to the loan_instance
        if 'document1' in request.FILES:
            loan_instance.document1 = document1
        if 'document2' in request.FILES:
            loan_instance.document2 = document2
        if 'document3' in request.FILES:
            loan_instance.document3 = document3
        if 'document4' in request.FILES:
            loan_instance.document4 = document4
        if 'document5' in request.FILES:
            loan_instance.document5 = document5
        loan_instance.save()

        return redirect('loan')  # Redirect to success page
    
    else:
        # If it's a GET request, initialize the form with existing data
        form = LoanEnquiryForm(instance=loan_instance)
        
        # Define document variables
        document1 = loan_instance.document1
        document2 = loan_instance.document2
        document3 = loan_instance.document3
        document4 = loan_instance.document4
        document5 = loan_instance.document5
    
    return render(request, 'loanedit.html', {'form': form, 'record_id': record_id,'document1': document1,
        'document2': document2,
        'document3': document3,
        'document4': document4,
        'document5': document5})

def update_loan(request, record_id):
    loan_instance = Loan.objects.get(pk=record_id)
    
    if request.method == 'POST':
        form = LoanEnquiryForm(request.POST, request.FILES, instance=loan_instance)
        if form.is_valid():
            form.save()
            return redirect('success_page')  # Replace 'success_page' with your actual URL name
        
    return render(request, 'loanedit.html', {'record_id': record_id, 'loan_instance': loan_instance})


def delete_loan_record(request, record_id):
    # Check if record exists
    record = get_object_or_404(Loan, pk=record_id)
    
    if request.method == 'POST':
    
    # Delete the record
        record.delete()
        return redirect('loan')  # Redirect to the loan records page
    return render(request, 'loan.html', {'record': record})

from django.http import JsonResponse
def get_record(request, record_id):
    record = get_object_or_404(Loan, pk=record_id)  # Replace LoanEnquiry with your actual model
    form = Loan(instance=record)  # Replace LoanEnquiry with your actual form
    return JsonResponse({'form': form})

def update_record(request, record_id):
    if request.method == 'POST':
        record = get_object_or_404(Loan, pk=record_id)  # Replace LoanEnquiry with your actual model
        form = Loan(request.POST, instance=record)  # Replace LoanEnquiry with your actual form
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
def policy_issue(request, record_id=None):
    # Existing code to fetch all data
    data = PolicyIssue.objects.all().order_by('-date')

    if request.method == 'POST':
        # Extracting form data
        date = request.POST.get('date')
        name = request.POST.get('name')
        number = request.POST.get('number')
        ckyc = request.POST.get('ckyc')
        p_number = request.POST.get('p_number')
        v_number = request.POST.get('v_number')
        Vehicle = request.POST.get('Vehicle')
        c_number = request.POST.get('c_number')
        e_number = request.POST.get('e_number')
        Location = request.POST.get('Location')
        HP_bank = request.POST.get('HP_bank')
        business_type = request.POST.get('business_type', 'Data')
        insurance_type = request.POST.get('insurance_type', 'TP')
        insurance_portal = request.POST.get('insurance_portal', 'Agency')
        I_company = request.POST.get('I_company')
        payment = request.POST.get('payment')
        payment_sos = request.POST.get('payment_sos')
        PS_date = request.POST.get('PS_date')
        PE_date = request.POST.get('PE_date')
        Ncb = request.POST.get('Ncb')
        Premium = request.POST.get('Premium')
        odNetPremium = request.POST.get('odNetPremium')
        commissionPercentage = request.POST.get('commissionPercentage')
        profitResult = request.POST.get('profitResult')
        tdsPercentage = request.POST.get('tdsPercentage', 5)
        profitAfterTDSResult = request.POST.get('profitAfterTDSResult')
        PayoutAmount = request.POST.get('PayoutAmount')
        payoutDiscount = request.POST.get('payoutDiscount')
        netProfitResult = request.POST.get('netProfitResult')
        Executive = request.POST.get('Executive')
        DSA = request.POST.get('DSA')

        # Check if the date is provided and not empty
        if date:
            try:
                # Convert the date string to a datetime object
                date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return HttpResponse('Invalid date format. Please use YYYY-MM-DD.')
        else:
            date = datetime.now()

        # If record_id is provided, update the existing record
        if record_id:
            policy_instance = PolicyIssue.objects.get(pk=record_id)
            policy_instance.date = date
            policy_instance.name = name
            policy_instance.number = number
            policy_instance.ckyc = ckyc
            policy_instance.p_number = p_number
            policy_instance.v_number = v_number
            policy_instance.Vehicle = Vehicle
            policy_instance.c_number = c_number
            policy_instance.e_number = e_number
            policy_instance.Location = Location
            policy_instance.HP_bank = HP_bank
            policy_instance.business_type = business_type
            policy_instance.insurance_type = insurance_type
            policy_instance.insurance_portal = insurance_portal
            policy_instance.I_company = I_company
            policy_instance.payment = payment
            policy_instance.payment_sos = payment_sos
            policy_instance.PS_date = PS_date
            policy_instance.PE_date = PE_date
            policy_instance.Ncb = Ncb
            policy_instance.Premium = Premium
            policy_instance.odNetPremium = odNetPremium
            policy_instance.commissionPercentage = commissionPercentage
            policy_instance.profitResult = profitResult
            policy_instance.tdsPercentage = tdsPercentage
            policy_instance.profitAfterTDSResult = profitAfterTDSResult
            policy_instance.PayoutAmount = PayoutAmount
            policy_instance.payoutDiscount = payoutDiscount
            policy_instance.netProfitResult = netProfitResult
            policy_instance.Executive = Executive
            policy_instance.DSA = DSA
            policy_instance.save()
            return redirect('success_page')  # Replace 'success_page' with your actual URL name

        else:
            # Create a new PolicyIssue instance
            PolicyIssue.objects.create(
                date=date,
                name=name,
                number=number,
                ckyc=ckyc,
                p_number=p_number,
                v_number=v_number,
                Vehicle=Vehicle,
                c_number=c_number,
                e_number=e_number,
                Location=Location,
                HP_bank=HP_bank,
                business_type=business_type,
                insurance_type=insurance_type,
                insurance_portal=insurance_portal,
                I_company=I_company,
                payment=payment,
                payment_sos=payment_sos,
                PS_date=PS_date,
                PE_date=PE_date,
                Ncb=Ncb,
                Premium=Premium,
                odNetPremium=odNetPremium,
                commissionPercentage=commissionPercentage,
                profitResult=profitResult,
                tdsPercentage=tdsPercentage,
                profitAfterTDSResult=profitAfterTDSResult,
                PayoutAmount=PayoutAmount,
                payoutDiscount=payoutDiscount,
                netProfitResult=netProfitResult,
                Executive=Executive,
                DSA=DSA
            )

            context = {"check": True}  # Assuming you have a check variable for success message
            return render(request, 'policyissue.html', context=context)

    # If record_id is provided, fetch the instance for editing
    if record_id:
        policy_instance = PolicyIssue.objects.get(pk=record_id)
        return render(request, 'policyissue.html', {'data': data, 'policy_instance': policy_instance})

    return render(request, 'policyissue.html', {'data': data})


from django.shortcuts import render, redirect
from datetime import datetime
from .models import PolicyIssue  # Import your PolicyIssue model

def edit_policy(request, record_id):
    try:
        record = PolicyIssue.objects.get(pk=record_id)
    except PolicyIssue.DoesNotExist:
        # Handle if record does not exist
        pass
    
    return render(request, 'editpolicy.html', {'record': record})

def update_policy(request, record_id):
    policy_instance = PolicyIssue.objects.get(pk=record_id)

    if request.method == 'POST':
        # Extracting form data
        date = request.POST.get('date')
        name = request.POST.get('name')
        number = request.POST.get('number')
        ckyc = request.POST.get('ckyc')
        p_number = request.POST.get('p_number')
        v_number = request.POST.get('v_number')
        Vehicle = request.POST.get('Vehicle')
        c_number = request.POST.get('c_number')
        e_number = request.POST.get('e_number')
        Location = request.POST.get('Location')
        HP_bank = request.POST.get('HP_bank')
        business_type = request.POST.get('business_type', 'Data')
        insurance_type = request.POST.get('insurance_type', 'TP')
        insurance_portal = request.POST.get('insurance_portal', 'Agency')
        I_company = request.POST.get('I_company')
        payment = request.POST.get('payment')
        payment_sos = request.POST.get('payment_sos')
        PS_date = request.POST.get('PS_date')
        PE_date = request.POST.get('PE_date')
        Ncb = request.POST.get('Ncb')
        Premium = request.POST.get('Premium')
        odNetPremium = request.POST.get('odNetPremium')
        commissionPercentage = request.POST.get('commissionPercentage')
        profitResult = request.POST.get('profitResult')
        tdsPercentage = request.POST.get('tdsPercentage', 5)
        profitAfterTDSResult = request.POST.get('profitAfterTDSResult')
        PayoutAmount = request.POST.get('PayoutAmount')
        payoutDiscount = request.POST.get('payoutDiscount')
        netProfitResult = request.POST.get('netProfitResult')
        Executive = request.POST.get('Executive')
        DSA = request.POST.get('DSA')

        # Check if the date is provided and not empty
        if date:
            try:
                # Convert the date string to a datetime object
                date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return HttpResponse('Invalid date format. Please use YYYY-MM-DD.')
        else:
            date = datetime.now()

        # Update the policy instance
        policy_instance.date = date
        policy_instance.name = name
        policy_instance.number = number
        policy_instance.ckyc = ckyc
        policy_instance.p_number = p_number
        policy_instance.v_number = v_number
        policy_instance.Vehicle = Vehicle
        policy_instance.c_number = c_number
        policy_instance.e_number = e_number
        policy_instance.Location = Location
        policy_instance.HP_bank = HP_bank
        policy_instance.business_type = business_type
        policy_instance.insurance_type = insurance_type
        policy_instance.insurance_portal = insurance_portal
        policy_instance.I_company = I_company
        policy_instance.payment = payment
        policy_instance.payment_sos = payment_sos
        policy_instance.PS_date = PS_date
        policy_instance.PE_date = PE_date
        policy_instance.Ncb = Ncb
        policy_instance.Premium = Premium
        policy_instance.odNetPremium = odNetPremium
        policy_instance.commissionPercentage = commissionPercentage
        policy_instance.profitResult = profitResult
        policy_instance.tdsPercentage = tdsPercentage
        policy_instance.profitAfterTDSResult = profitAfterTDSResult
        policy_instance.PayoutAmount = PayoutAmount
        policy_instance.payoutDiscount = payoutDiscount
        policy_instance.netProfitResult = netProfitResult
        policy_instance.Executive = Executive
        policy_instance.DSA = DSA

        # Save the updated instance
        policy_instance.save()

        return redirect('policy_issue')  # Replace 'policy_issue' with your actual URL name


def delete_record(request, record_id):
    # Retrieve the record or return 404 error if not found
    record = get_object_or_404(PolicyIssue, pk=record_id)
    
    if request.method == 'POST':
        # If the request method is POST, it means the user has confirmed deletion
        record.delete()  # Delete the record
        return redirect('policy_issue')  # Redirect to the list page
        
    # If the request method is GET, display a confirmation page
    return render(request, 'policy.html', {'record': record})
                  
import pandas as pd
from django.shortcuts import render
from .models import PolicyIssue
from openpyxl import load_workbook
import os

def upload_policy(request):
    if request.method == 'POST':
        xlsx_file = request.FILES.get('XLSPolicy')

        if not xlsx_file:
            return HttpResponse('No file uploaded')

        try:
            wb = load_workbook(xlsx_file, read_only=True)
            ws = wb.active
            
            # Assuming the first row contains headers
            headers = [cell.value for cell in ws[1]]
            
            # Mapping headers to column indices
            header_indices = {header: index for index, header in enumerate(headers)}
            
            # Required columns
            required_columns = ['date', 'name', 'number','ckyc' , 'p_number', 'v_number', 'Vehicle', 
                                'c_number', 'e_number', 'Location', 'HP_bank', 'business_type', 
                                'insurance_type', 'insurance_portal', 'I_company', 'payment', 
                                'payment_sos', 'PS_date', 'PE_date', 'Ncb', 'Premium', 
                                'odNetPremium', 'commissionPercentage', 'profitResult', 'tdsPercentage', 
                                'profitAfterTDSResult','payoutDiscount','PayoutAmount', 'netProfitResult', 
                                'Executive', 'DSA']
            
            for column in required_columns:
                if column not in header_indices:
                    return HttpResponse(f'Error: Column "{column}" not found in the Excel file')
            
            for row in ws.iter_rows(min_row=2):
                # Initialize a dictionary to store cell values
                row_data = {}
                for header, index in header_indices.items():
                    # Extract cell value
                    cell_value = row[index].value
                    # Check if cell value is a formula or expression
                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        # If it's a formula, set it to None for now
                        cell_value = None
                    row_data[header] = cell_value
                
                # Validate and correct date formats
                for date_field in ['date', 'PS_date', 'PE_date']:
                    if row_data.get(date_field) and isinstance(row_data[date_field], str):
                        row_data[date_field] = parse_date(row_data[date_field])

                # Create PolicyIssue instance
                policy_issue = PolicyIssue(**row_data)
                policy_issue.save()

            return HttpResponse('Data uploaded successfully!')

        except Exception as e:
            return HttpResponse(f'Error uploading data: {e}')

    return render(request, 'uploadxlspolicyissue.html')

def parse_date(date_str):
    formats = ['%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y']  # Add more formats as needed
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    #raise ValueError('Invalid date format')
    

    
from django.shortcuts import render, redirect
from .models import Loan

from datetime import datetime
from django.shortcuts import render, HttpResponse
from .models import PolicyIssue

def renewal(request):
    data = PolicyIssue.objects.all().order_by('-date')
    context = {}

    if request.method == 'POST':
        end_date = request.POST.get('end_date')
        start_date = request.POST.get('start_date')

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                data = data.filter(PE_date__lte=end_date)
            except ValueError:
                return HttpResponse('Invalid end date format. Please use YYYY-MM-DD.')

        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                if end_date:
                    data = data.filter(PE_date__gte=start_date)
                    total_dates = (end_date - start_date).days + 1
                    context['total_dates'] = total_dates
            except ValueError:
                return HttpResponse('Invalid start date format. Please use YYYY-MM-DD.')

        if not end_date and not start_date:
            # Extracting form data
            date = request.POST.get('date')
            name = request.POST.get('name')
            number = request.POST.get('number')
            p_number = request.POST.get('p_number')
            v_number = request.POST.get('v_number')
            Vehicle = request.POST.get('Vehicle')
            c_number = request.POST.get('c_number')
            e_number = request.POST.get('e_number')
            Location = request.POST.get('Location')
            HP_bank = request.POST.get('HP_bank')
            business_type = request.POST.get('business_type', 'Data')
            insurance_type = request.POST.get('insurance_type', 'TP')
            insurance_portal = request.POST.get('insurance_portal', 'Agency')
            I_company = request.POST.get('I_company')
            payment = request.POST.get('payment')
            payment_sos = request.POST.get('payment_sos')
            PS_date = request.POST.get('PS_date')
            PE_date = request.POST.get('PE_date')
            Ncb = request.POST.get('Ncb')
            Premium = request.POST.get('Premium')
            odNetPremium = request.POST.get('odNetPremium')
            commissionPercentage = request.POST.get('commissionPercentage')
            profitResult = request.POST.get('profitResult')
            tdsPercentage = request.POST.get('tdsPercentage', 5)
            profitAfterTDSResult = request.POST.get('profitAfterTDSResult')
            PayoutAmount = request.POST.get('PayoutAmount')
            payoutDiscount = request.POST.get('payoutDiscount')
            netProfitResult = request.POST.get('netProfitResult')
            Executive = request.POST.get('Executive')
            DSA = request.POST.get('DSA')

            # Check if the date is provided and not empty
            if date:
                try:
                    # Convert the date string to a datetime object
                    date = datetime.strptime(date, '%Y-%m-%d').date()
                except ValueError:
                    return HttpResponse('Invalid date format. Please use YYYY-MM-DD.')
            else:
                date = datetime.now().date()  # Assigning default value if date is not provided

            # Create a new PolicyIssue instance
            PolicyIssue.objects.create(
                date=date,
                name=name,
                number=number,
                p_number=p_number,
                v_number=v_number,
                Vehicle=Vehicle,
                c_number=c_number,
                e_number=e_number,
                Location=Location,
                HP_bank=HP_bank,
                business_type=business_type,
                insurance_type=insurance_type,
                insurance_portal=insurance_portal,
                I_company=I_company,
                payment=payment,
                payment_sos=payment_sos,
                PS_date=PS_date,
                PE_date=PE_date,
                Ncb=Ncb,
                Premium=Premium,
                odNetPremium=odNetPremium,
                commissionPercentage=commissionPercentage,
                profitResult=profitResult,
                tdsPercentage=tdsPercentage,
                profitAfterTDSResult=profitAfterTDSResult,
                PayoutAmount=PayoutAmount,
                payoutDiscount=payoutDiscount,
                netProfitResult=netProfitResult,
                Executive=Executive,
                DSA=DSA
            )

            context['check'] = True  # Assuming you have a check variable for success message

    return render(request, 'renewal.html', {'data': data, **context})

import csv
from django.http import HttpResponse
from .models import PolicyIssue

def download_data(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="data.csv"'

    writer = csv.writer(response)
    writer.writerow(['Field1', 'Field2', 'Field3'])  # Add headers

    queryset = PolicyIssue.objects.all()  # Fetch all data
    for obj in queryset:
        writer.writerow([obj.field1, obj.field2, obj.field3])  # Add rows

    return response


from django.http import JsonResponse
from .models import PolicyIssue

def get_data(request):
    data = PolicyIssue.objects.all().values()  # Fetch all data from YourModel
    return JsonResponse(list(data), safe=False)

from django.db.models import Sum
from django.db.models.functions import ExtractYear
from .models import PolicyIssue
from django.db.models import Avg
import json
from django.shortcuts import render


# def chart(request):
#     # Retrieve data for Commission Percentage
#     avg_commission = PolicyIssue.objects.aggregate(avg_commission=Avg('commissionPercentage'))['avg_commission']

#     # Retrieve data for Premium and Insurance Company
#     premium_data = PolicyIssue.objects.values('I_company').annotate(total_premium=Sum('Premium')).order_by('-total_premium')[:5]
#     premium_data_list = [{'company': data['I_company'], 'premium': float(data['total_premium'])} for data in premium_data]

#     # Retrieve data for Net Profit by Business Type
#     profit_data = PolicyIssue.objects.values('business_type').annotate(net_profit=Sum('netProfitResult'))
#     profit_data_list = [{'business_type': data['business_type'], 'net_profit': float(data['net_profit'])} for data in profit_data]

#     # Retrieve data for Net Profit by Insurance Type (for doughnut chart)
#     net_profit_by_insurance = PolicyIssue.objects.values('insurance_type').annotate(net_profit=Sum('netProfitResult'))
#     doughnut_data_list = [{'insurance_type': data['insurance_type'], 'net_profit': float(data['net_profit'])} for data in net_profit_by_insurance]

#     # Retrieve data for Line Chart (Net Profit by Year, OD/Net Premium by Year, Profit After TDS by Year)
#     line_data = PolicyIssue.objects.annotate(
#         month=Extractmonth('date')
#     ).values('month').annotate(
#         net_profit=Sum('netProfitResult'), 
#         od_net_premium=Sum('odNetPremium'),  
#         profit_after_tds=Sum('profitAfterTDSResult')
#     )

#    # Convert line data to list of dictionaries
#     line_data_list = [
#         {
#             'month': str(item['month']),  # Ensure 'year' is converted to a string
#             'net_profit': float(item['net_profit']),
#             'od_net_premium': float(item['od_net_premium']),
#             'profit_after_tds': float(item['profit_after_tds'])
#         }
#         for item in line_data
#     ]

#     # Serialize the data to JSON format
#     premium_data_json = json.dumps(premium_data_list)
#     profit_data_json = json.dumps(profit_data_list)
#     doughnut_data_json = json.dumps(doughnut_data_list)
#     line_data_json = json.dumps(line_data_list)

#     return render(request, 'chart.html', {
#         'avg_commission': avg_commission,
#         'premium_data_json': premium_data_json,
#         'profit_data_json': profit_data_json,
#         'doughnut_data_json': doughnut_data_json,
#         'line_data_json': line_data_json
#     })


from django.shortcuts import render
from django.http import HttpResponse
import pymysql
import pandas as pd

def fetch_data(request):
    # Connect to the MySQL database
    conn = pymysql.connect(host=' db2.crqssyqccgp1.eu-north-1.rds.amazonaws.com', user='nrvkakkad', password='Kishan#32', database='db2')

    # Read data from a table into a DataFrame
    df = pd.read_sql_query("SELECT * FROM blog_policyissue", conn)
    conn.close()

    # Return the DataFrame
    return df
# import plotly.express as px
# import plotly.graph_objects as go
# from django.shortcuts import render

# def chart(request):
#     # Fetching data from the DataFrame
#     df = fetch_data(request)
    
#     # Convert 'date' column to datetime format
#     df['date'] = pd.to_datetime(df['date'])

#     # Grouping data by insurance company and summing premiums
#     df_grouped = df.groupby('I_company')['Premium'].sum().reset_index()

#     # Sorting the companies based on total premiums
#     df_sorted = df_grouped.sort_values(by='Premium', ascending=False)

#     # Selecting the top 5 companies
#     top_5_companies = df_sorted.head(5)

#     # Plotting the bar chart
#     fig_bar = px.bar(top_5_companies, x="I_company", y="Premium", title="Top 5 Insurance Companies by Premium")

#     fig_bar.update_layout(
#     # autosize=False,
#     width=500,
#     height=500,
#     margin=dict(
#         l=50,
#         r=50,
#         b=100,
#         t=100,
#         pad=4
#     ),
#     plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
#     paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
# )

#     # Extracting year from the date column
#     df['Year'] = df['date'].dt.year.astype(str)

#     df1 = df.groupby(by="Year")[["netProfitResult", "profitAfterTDSResult", "odNetPremium"]].sum().rename(columns={"netProfitResult": "Net Profit", "profitAfterTDSResult": "Profit After TDS", "odNetPremium": "OD Net Premium"})

#     # Plotting the line chart with only the year on the x-axis
#     fig_line = px.line(df1, x=["Year"], y=["Total Premium"], 
#                    title='Net Profit Over Time with Additional Metrics')
    
#     fig_line.update_layout(
#     # autosize=False,
#     width=500,
#     height=500,
#     margin=dict(
#         l=50,
#         r=50,
#         b=100,
#         t=100,
#         pad=4
#     ),
#     plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
#     paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
# )

#     # Plotting the donut chart using Plotly Graph Objects
    
#     fig_donut_2 = go.Figure(data=[go.Pie(labels=df['business_type'], values=df['netProfitResult'], hole=0.5)])
#     fig_donut_2.update_layout(title="Net profit by insurance type")

#     fig_donut_2.update_layout(
#     # autosize=False,
#     width=500,
#     height=500,
#     margin=dict(
#         l=50,
#         r=50,
#         b=100,
#         t=100,
#         pad=4
#     ),
#     plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
#     paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
# )
    

#     # Plotting the donut chart using Plotly Graph Objects
#     fig_donut = go.Figure(data=[go.Pie(labels=df['insurance_type'], values=df['netProfitResult'], hole=0.5)])
#     fig_donut.update_layout(title="Net profit by insurance type")

#     fig_donut.update_layout(
#     # autosize=False,
#     width=500,
#     height=500,
#     margin=dict(
#         l=50,
#         r=50,
#         b=100,
#         t=100,
#         pad=4
#     ),
#     plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
#     paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
# )
#     # fig_funnel = px.funnel(df, x='Location', y='Premium')
#     df_grouped = df.groupby('Location')['Premium'].sum().reset_index()
#     df_sorted = df_grouped.sort_values(by='Premium', ascending=False)

#     # Selecting the top 5 companies
#     top_5_companies = df_sorted.head(5)

#     # Plotting the bar chart
#     fig_funnel = px.funnel(top_5_companies, x="Premium", y="Location", title="Location by Premium",orientation='h')

#     fig_funnel.update_layout(
#     # autosize=False,
#     width=500,
#     height=500,
#     margin=dict(
#         l=50,
#         r=50,
#         b=100,
#         t=100,
#         pad=4
#     ),
#     plot_bgcolor='rgba(0,0,0,0)',  # Transparent background
#     paper_bgcolor='rgba(0,0,0,0)',  # Transparent plot area background
# )

#     # Convert the Plotly figures to HTML
#     plot_html_bar = fig_bar.to_html(full_html=False)
#     plot_html_line = fig_line.to_html(full_html=False)
#     plot_html_donut_2 = fig_donut_2.to_html(full_html=False)
#     plot_html_donut = fig_donut.to_html(full_html=False)
#     plot_html_funnel = fig_funnel.to_html(full_html=False)

#     # Pass the HTML to the template for rendering
    
#     return render(request, 'chart.html', {'plot_html_bar': plot_html_bar, 'plot_html_line': plot_html_line, 'plot_html_donut_2': plot_html_donut_2, 'plot_html_donut': plot_html_donut,'plot_html_funnel':plot_html_funnel})


# from django.shortcuts import render
# import plotly.graph_objects as go

# from django.db.models import Sum
# from django.shortcuts import render
# import plotly.graph_objects as go
# from .models import PolicyIssue  # Import your model here

# def dashboard(request):
#     # Retrieve data from your database
#     queryset = PolicyIssue.objects.all()  # Replace PolicyIssue with your actual model

#     # Calculate metrics (example calculation)
#        # Calculate metrics
#     total_p_number = queryset.count()
#     total_premium = queryset.aggregate(total_premium=Sum('Premium'))['total_premium']
#     total_odnet_premium = queryset.aggregate(total_odnet_premium=Sum('odNetPremium'))['total_odnet_premium']
#     total_commission_percentage = queryset.aggregate(total_commission_percentage=Sum('commissionPercentage'))['total_commission_percentage']
#     total_profit = queryset.aggregate(total_profit=Sum('profitResult'))['total_profit']
#     total_net_profit = queryset.aggregate(total_net_profit=Sum('netProfitResult'))['total_net_profit']
#     total_profit_after_tds = queryset.aggregate(total_profit_after_tds=Sum('profitAfterTDSResult'))['total_profit_after_tds']

#     # Create Plotly figure for each metric

#     fig_profit = go.Figure(go.Indicator(
#         mode="number",
#         value=total_profit,
#         title="Total Profit"
#     ))

#     fig_net_profit = go.Figure(go.Indicator(
#         mode="number",
#         value=total_net_profit,
#         title="Total Net Profit"
#     ))

#     fig_profit_after_tds = go.Figure(go.Indicator(
#         mode="number",
#         value=total_profit_after_tds,
#         title="Total Profit After TDS"
#     ))

#     # Create Plotly figures for each metric
#     fig_customer = go.Figure(go.Indicator(
#         mode="number",
#         value=total_p_number,
#         title="Total Customer"
#     ))

#     fig_premium = go.Figure(go.Indicator(
#         mode="number",
#         value=total_premium,
#         title="Total Premium"
#     ))

#     fig_odnet_premium = go.Figure(go.Indicator(
#         mode="number",
#         value=total_odnet_premium,
#         title="Total OD/Net Premium"
#     ))

#     fig_commission_percentage = go.Figure(go.Indicator(
#         mode="number",
#         value=total_commission_percentage,
#         title="Total Commission"
#     ))

#     # Convert Plotly figures to HTML
#     plot_html_profit = fig_profit.to_html(full_html=False)
#     plot_html_net_profit = fig_net_profit.to_html(full_html=False)
#     plot_html_profit_after_tds = fig_profit_after_tds.to_html(full_html=False)
#     plot_html_p_number = fig_customer.to_html(full_html=False)
#     plot_html_premium = fig_premium.to_html(full_html=False)
#     plot_html_odnet_premium = fig_odnet_premium.to_html(full_html=False)
#     plot_html_commission_percentage = fig_commission_percentage.to_html(full_html=False)

#        # Pass the HTML to the template for rendering
#     return render(request, 'dashboard.html', {'plot_html_profit': plot_html_profit,
#                                               'plot_html_net_profit': plot_html_net_profit,
#                                               'plot_html_profit_after_tds': plot_html_profit_after_tds,
#                                               'plot_html_p_number': plot_html_p_number,
#                                               'plot_html_premium': plot_html_premium,
#                                               'plot_html_odnet_premium': plot_html_odnet_premium,
#                                               'plot_html_commission_percentage': plot_html_commission_percentage})


import plotly.graph_objects as go
from django.shortcuts import render
from .models import Loan  # Import your Loan model here

def loan_dashboard(request):
    # Retrieve data from your Loan model
    queryset = Loan.objects.all()

    # Calculate the total number of loan inquiries
    total_name = queryset.count()

    # Create Plotly figures for each metric
    fig_total_name = go.Figure(go.Indicator(
        mode="number",
        value=total_name,
        title="Total Customer"
    ))

    plot_html_total_name = fig_total_name.to_html(full_html=False)

    # Pass the total number of inquiries to the template for rendering
    return render(request, 'loan_dashboard.html', {'plot_html_total_name': plot_html_total_name})


from django.shortcuts import render
from .models import InsuranceEnquiry
import plotly.graph_objects as go

from django.shortcuts import render
from .models import InsuranceEnquiry
import plotly.graph_objects as go

def insurance_dashboard(request):
    # Retrieve data from your InsuranceEnquiry model
    inquiries = InsuranceEnquiry.objects.all()

    # Get the total count of unique names
    total_unique_names = inquiries.values('name').distinct().count()

    # Create a Plotly figure for the total count of unique names
    fig_total_unique_names = go.Figure(go.Indicator(
        mode="number",
        value=total_unique_names,
        title="Total Unique Names",
    ))

    # Convert Plotly figure to HTML
    plot_html_total_unique_names = fig_total_unique_names.to_html(full_html=False)

    # Pass the Plotly HTML to the template for rendering
    return render(request, 'insurance_dashboard.html', {'plot_html_total_unique_names': plot_html_total_unique_names})


